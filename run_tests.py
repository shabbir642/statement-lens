#!/usr/bin/env python3
"""End-to-end tests. Reports what actually happened, including failures.

Run: .venv/bin/python run_tests.py
"""

import offline_guard  # noqa: F401  — prove the pipeline runs with sockets blocked

import json
import os
import subprocess
import sys
import tempfile

import make_test_pdf
from extractor import extract_pdf, reconcile
from categorise import ensure_categories_file, categorise
from reporter import build_report
from analysis import load_transactions

RESULTS = []

# Scratch dir for test artefacts (CSV + HTML), so a run never touches the
# repo's real transactions.csv / report.html.  Removed on exit.
TMPDIR = tempfile.mkdtemp(prefix="lens-test-")


def check(name, ok, detail=""):
    RESULTS.append((name, ok, detail))
    mark = "PASS" if ok else "FAIL"
    print(f"[{mark}] {name}" + (f"  — {detail}" if detail else ""))
    return ok


def main():
    categories = ensure_categories_file()

    # 1. Generate a realistic statement.
    opening, closing, nrows = make_test_pdf.generate()
    print(f"generated test PDF: opening={opening:.2f} closing={closing:.2f} "
          f"rows={nrows}")

    # 2. Extract and check exact reconciliation.
    res = extract_pdf("statements/test-hdfc.pdf")
    rows = res["rows"]
    parsed_sum = sum(r.signed for r in rows if r.signed is not None)
    recomputed_closing = opening + parsed_sum
    check("parsed row count is plausible",
          abs(len(rows) - nrows) <= 1,
          f"parsed {len(rows)} of {nrows} generated")
    check("opening + sum(parsed) == closing, to the paisa",
          round(recomputed_closing, 2) == round(closing, 2),
          f"{recomputed_closing:.2f} vs {closing:.2f}")
    check("reconciliation self-check reports OK",
          res["reconcile"]["ok"] is True,
          res["reconcile"]["message"])
    check("every row parsed at high confidence (balance column trusted)",
          all(r.confidence == "high" for r in rows),
          f"{sum(1 for r in rows if r.confidence=='high')}/{len(rows)} high")

    # 3. Negative test: drop one row, reconciliation MUST now report a gap.
    dropped = rows[:len(rows)//2] + rows[len(rows)//2 + 1:]
    neg = reconcile(dropped, res["opening"], res["closing"])
    check("dropping one row breaks reconciliation (non-zero gap)",
          neg[0] is False and abs(neg[4]) > 0.01,
          neg[1])

    # 4. Salary must categorise as Income, not Transfers Out.
    salary = [r for r in rows if "SALARY" in r.description.upper()]
    cats = {categorise(r.description, r.signed, categories)[0] for r in salary}
    check("salary credits categorise as Income",
          bool(salary) and cats == {"Income"},
          f"{len(salary)} salary rows -> {cats}")

    # 4b. The CREDIT trap: description not corrupted, parsed as a credit.
    interest = [r for r in rows if "INTEREST CAPITALISED" in r.description.upper()]
    check("CREDIT INTEREST line survives (description intact, parsed as credit)",
          bool(interest)
          and all(r.signed and r.signed > 0 for r in interest)
          and all(r.description.upper() == "CREDIT INTEREST CAPITALISED"
                  for r in interest),
          f"{len(interest)} interest rows, sample: "
          f"{interest[0].description if interest else 'none'}")

    # 4c. Merchant normalisation drops the single-letter token ("D").
    from categorise import normalise_merchant
    m = normalise_merchant("ACH D- HDFC MUTUAL FUND SIP")
    check("merchant normalisation drops single-letter token",
          "HDFC MUTUAL FUND" in m and " D " not in f" {m} ", f"-> '{m}'")

    # 5. Build the report and check offline-safety + populated sections.
    #    Write to a throwaway temp dir, never the repo's own transactions.csv /
    #    report.html — those hold the user's real, gitignored, unrecoverable
    #    financial history, and a test run must never clobber them.
    from lens import _write_csv
    test_csv = os.path.join(TMPDIR, "transactions.csv")
    test_report = os.path.join(TMPDIR, "report.html")
    _write_csv(rows, test_csv)
    out, n = build_report(test_csv, categories, out_path=test_report,
                          reconcile=res["reconcile"])
    with open(out, encoding="utf-8") as fh:
        html = fh.read()
    check("report.html has zero 'http' references (offline-safe)",
          "http" not in html, f"grep http -> {html.count('http')} hits")

    # 6. Store merge: re-import is a no-op; distinct same-key rows both survive.
    import store
    rec = [{"date": "2026-06-01", "description": "X",
            "amount": "-100.00", "confidence": "high"}]
    _, added, skipped = store.merge(rec, rec)
    check("store: re-importing a statement adds nothing",
          added == 0 and skipped == 1, f"added={added} skipped={skipped}")
    _, added2, _ = store.merge([], rec + rec)
    check("store: two genuine same-key rows both kept", added2 == 2,
          f"added={added2}")

    # 7. Self-transfer netting: matched debit+credit pair tagged, small ones left.
    from analysis import tag_internal_transfers
    tt = [
        {"date": "2026-03-01", "amount": -5000.0, "internal": False},
        {"date": "2026-03-02", "amount": 5000.0, "internal": False},
        {"date": "2026-03-01", "amount": -50.0, "internal": False},  # below floor
    ]
    pairs = tag_internal_transfers(tt)
    check("self-transfer netting matches a debit+credit pair",
          pairs == 1 and tt[0]["internal"] and tt[1]["internal"]
          and not tt[2]["internal"], f"pairs={pairs}")

    # 8. Alerts: net-negative fires; balance timing-risk fires as 'high'.
    from alerts import spending_alerts, balance_alerts
    sa = spending_alerts([
        {"date": "2026-01-01", "month": "2026-01", "amount": -10000.0,
         "category": "UPI Payment", "confidence": "high", "internal": False},
        {"date": "2026-01-02", "month": "2026-01", "amount": 100.0,
         "category": "Income", "confidence": "high", "internal": False},
    ])
    check("alerts: net-negative alert fires",
          any("Net negative" in a["msg"] for a in sa))
    ba = balance_alerts([{"low_balance": 800, "low_date": "2026-01-05",
                          "big_debit_amt": 20000, "big_debit_date": "2026-01-08",
                          "low_before_big": True, "churn": 3}])
    check("alerts: low-balance-before-big-debit is a high alert",
          any(a["level"] == "high" for a in ba))

    # 9. Per-bank golden test: synthetic SBI layout (two accounts, wrapped
    #    rows, null-padded balance lines) must parse cleanly.
    test_sbi_layout(categories)

    # 10. Tabular ingest: CSV statements map columns and reconcile like PDFs.
    test_tabular_csv()

    render_headlessly(out)

    print("\n" + "=" * 60)
    passed = sum(1 for _, ok, _ in RESULTS if ok)
    print(f"{passed}/{len(RESULTS)} checks passed")
    return 0 if passed == len(RESULTS) else 1


def _cleanup():
    import shutil
    shutil.rmtree(TMPDIR, ignore_errors=True)


def test_tabular_csv():
    """CSV ingest: detect columns past a preamble, sign Dr/Cr, reconcile, and
    pass our own reviewed-CSV schema through untouched."""
    from extractor import extract

    # A realistic bank export: preamble rows, Dr/Cr split, a balance column, a
    # trailing summary row — none of which should become transactions.
    bank = os.path.join(TMPDIR, "bank.csv")
    with open(bank, "w", encoding="utf-8") as fh:
        fh.write("Account Statement for A/C 001234567890\n\n")
        fh.write("Txn Date,Value Date,Narration,Cheque No,"
                 "Withdrawal (Dr),Deposit (Cr),Closing Balance\n")
        fh.write('02/06/2024,02/06/2024,UPI-ZOMATO-ORDER,,"1,200.00",,"48,800.00"\n')
        fh.write('05/06/2024,05/06/2024,NEFT-ACME-SALARY,,,"50,000.00","98,800.00"\n')
        fh.write('09/06/2024,09/06/2024,ATM-CASH WDL,,"5,000.00",,"93,800.00"\n')
        fh.write('30/06/2024,,Closing Balance,,,,"93,800.00"\n')
    res = extract(bank)
    rows = res["rows"]
    signed = {r.description: r.signed for r in rows}
    check("CSV: header found past preamble, summary row dropped (3 txns)",
          len(rows) == 3, f"parsed {len(rows)}")
    check("CSV: Dr/Cr split becomes signed amounts",
          signed.get("UPI-ZOMATO-ORDER") == -1200.0
          and signed.get("NEFT-ACME-SALARY") == 50000.0,
          f"{signed}")
    check("CSV: reconciles against the balance column",
          res["reconcile"]["ok"] is True, res["reconcile"]["message"])

    # No balance column -> exact amounts, but reconciliation is skipped (not a
    # failure), and the message says so.
    nobal = os.path.join(TMPDIR, "nobal.csv")
    with open(nobal, "w", encoding="utf-8") as fh:
        fh.write("Date;Description;Amount\n01-07-2024;Netflix;-199.00\n")
    r2 = extract(nobal)
    check("CSV: no-balance file parses (semicolon), reconciliation skipped",
          r2["reconcile"]["ok"] is None and len(r2["rows"]) == 1
          and r2["rows"][0].signed == -199.0,
          r2["reconcile"]["message"][:50])

    # Our own reviewed CSV schema is passed through verbatim (no re-detection).
    intern = os.path.join(TMPDIR, "reviewed.csv")
    with open(intern, "w", encoding="utf-8") as fh:
        fh.write("date,description,amount,confidence\n"
                 "2026-07-07,SALARY,181684.00,high\n")
    r3 = extract(intern)
    check("CSV: our reviewed-CSV schema passes through unchanged",
          len(r3["rows"]) == 1 and r3["rows"][0].signed == 181684.0
          and r3["rows"][0].confidence == "high",
          f"{[(x.date, x.signed) for x in r3['rows']]}")

    # Bare "Dr"/"Cr" column headers (some banks) must be recognised, while a
    # column like "Address" (contains the substring "dr") must NOT be.
    from tabular_source import _classify
    check("CSV: bare 'Dr'/'Cr' headers classify; 'Address'/'Order' do not",
          _classify("Dr") == "debit" and _classify("Cr") == "credit"
          and _classify("Dr (₹)") == "debit"
          and _classify("Address") is None and _classify("Order No") is None,
          f"Dr={_classify('Dr')} Cr={_classify('Cr')} Address={_classify('Address')}")
    bare = os.path.join(TMPDIR, "bare.csv")
    with open(bare, "w", encoding="utf-8") as fh:
        fh.write("Date,Particulars,Dr,Cr,Balance\n")
        fh.write("02/06/2024,SHOP,300.00,,700.00\n")
        fh.write("03/06/2024,REFUND,,500.00,1200.00\n")
    rb = extract(bare)
    sb = {r.description: r.signed for r in rb["rows"]}
    check("CSV: bare Dr/Cr columns sign correctly and reconcile",
          sb.get("SHOP") == -300.0 and sb.get("REFUND") == 500.0
          and rb["reconcile"]["ok"] is True, f"{sb} {rb['reconcile']['ok']}")

    test_tabular_xlsx()


def test_tabular_xlsx():
    """XLSX ingest: skip a cover sheet, parse typed/serial/string dates, and
    reconcile — mirroring the CSV path via the same shared core."""
    try:
        import openpyxl
    except ImportError:
        check("XLSX: openpyxl available", None,
              "SKIPPED — pip install openpyxl to enable the Excel path")
        return
    import datetime
    from extractor import extract

    path = os.path.join(TMPDIR, "bank.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["My Bank - Account Statement"])          # preamble
    ws.append([])
    ws.append(["Txn Date", "Narration", "Debit", "Credit", "Balance"])
    ws.append([datetime.datetime(2024, 6, 2), "UPI-ZOMATO", 1200.00, None, 48800.00])
    ws.append([45448, "NEFT-SALARY", None, 50000.00, 98800.00])   # serial date
    ws.append(["09/06/2024", "ATM-WDL", "5,000.00", None, "93,800.00"])
    ws.append(["Total", None, None, None, None])         # trailer
    cover = wb.create_sheet("Cover")
    cover.append(["Statement summary"])
    wb.move_sheet("Cover", -(len(wb.sheetnames) - 1))    # cover first — must be skipped
    wb.save(path)

    res = extract(path)
    rows = res["rows"]
    dates = [r.date for r in rows]
    signed = {r.description: r.signed for r in rows}
    check("XLSX: cover sheet skipped, 3 txns from the real sheet",
          len(rows) == 3, f"parsed {len(rows)}: {dates}")
    check("XLSX: typed, serial and string dates all parse",
          dates == ["2024-06-02", "2024-06-05", "2024-06-09"], f"{dates}")
    check("XLSX: Dr/Cr split signs correctly and reconciles",
          signed.get("UPI-ZOMATO") == -1200.0
          and signed.get("NEFT-SALARY") == 50000.0
          and res["reconcile"]["ok"] is True,
          res["reconcile"]["message"])

    # A file renamed to .xlsx that isn't a real workbook must fail with a clear
    # message, not a raw BadZipFile/openpyxl traceback.
    from tabular_source import TabularParseError
    fake = os.path.join(TMPDIR, "fake.xlsx")
    with open(fake, "w", encoding="utf-8") as fh:
        fh.write("Date,Description,Amount\n01-07-2024,X,-1.00\n")  # really CSV
    try:
        extract(fake)
        ok = False
    except TabularParseError:
        ok = True
    except Exception:
        ok = False
    check("XLSX: a mislabeled/invalid .xlsx fails with a clear message", ok)


def test_sbi_layout(categories):
    """Golden test for the SBI-specific messes (regression guard)."""
    import make_sbi_pdf
    from analysis import balance_health
    exp = make_sbi_pdf.generate()
    res = extract_pdf(exp["path"])
    rows = res["rows"]

    check("SBI: row count matches (no wrapped rows lost, no null garbage)",
          len(rows) == exp["total"], f"parsed {len(rows)} of {exp['total']}")
    check("SBI: no balance-summary/null line ingested as a transaction",
          all("null" not in (r.description or "").lower()
              and "oupllening" not in (r.description or "").lower()
              for r in rows),
          "clean" if rows else "no rows")
    check("SBI: wrapped narration recovered (not dropped)",
          any(exp["wrap_token"] in (r.description or "").lower() for r in rows),
          f"looked for '{exp['wrap_token']}'")
    check("SBI: both accounts detected",
          len(balance_health(rows)) == 2,
          f"{len(balance_health(rows))} accounts")
    check("SBI: reconciles across accounts and pages",
          res["reconcile"]["ok"] is True, res["reconcile"]["message"])


def render_headlessly(report_path):
    """Render via render_check.py in a clean (guard-free) subprocess.

    Playwright talks to Chromium over a local socket, so it cannot run in this
    offline-guarded process; the subprocess only reads the local HTML file.
    """
    proc = subprocess.run(
        [sys.executable, "render_check.py", report_path],
        capture_output=True, text=True)
    try:
        data = json.loads(proc.stdout.strip().splitlines()[-1])
    except Exception:
        check("headless render (playwright)", False,
              f"could not run render check: {proc.stderr[-300:]}")
        return
    if not data.get("available"):
        check("headless render (playwright)", None,
              "SKIPPED — playwright not installed; `pip install playwright` "
              "then `playwright install chromium` to enable")
        return
    if data.get("crash"):
        check("headless render (playwright)", False,
              f"render crashed: {data['crash']}")
        return
    check("headless render: zero console errors",
          len(data["errors"]) == 0, f"errors={data['errors'][:3]}")
    empty = [s for s, v in data["populated"].items() if not v]
    check("headless render: every section populated",
          not empty, f"empty sections: {empty}" if empty else "all populated")


if __name__ == "__main__":
    try:
        code = main()
    finally:
        _cleanup()
    sys.exit(code)
